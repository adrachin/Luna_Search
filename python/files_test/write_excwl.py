import openpyxl
from urllib.parse import quote
import os

def validate_path(path):
    return os.path.exists(path)

# Pfad zur Textdatei
text_file_path = '/Users/thomas/_python/MyApps/Luna_Search/templates/treffer.txt'

# Neues Excel-Workbook erstellen
wb = openpyxl.Workbook()
ws = wb.active

# Spaltenüberschriften und Formatierung
headers = ['Projekt Name', 'SSD', 'Öffnen', 'Öffnen in Luna']
ws.append(headers)
for col in range(1, 5):
    ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=1, column=col).fill = openpyxl.styles.PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

# Textdatei öffnen und jede Zeile verarbeiten
with open(text_file_path, 'r', encoding='utf-8') as file:
    for index, line in enumerate(file, start=2):
        line = line.strip()
        if validate_path(line):
            volume = line.split('/')[2]  # Extrahiert den Volumennamen
            project_name = line.split('/')[-1]  # Extrahiert den Projektname

            # Link, um den Pfad im Finder zu öffnen
            finder_link = f"file://{quote(line)}"
            
            # Link, um das Projekt in Luna zu öffnen
            luna_link = f"luna://{quote(line)}"

            # Daten in die Zeilen schreiben
            ws.append([project_name, volume, 'Im Finder öffnen', 'In Luna öffnen'])
            ws[f'C{index}'].hyperlink = finder_link
            ws[f'C{index}'].style = 'Hyperlink'
            ws[f'D{index}'].hyperlink = luna_link
            ws[f'D{index}'].style = 'Hyperlink'
        else:
            print(f"Pfad nicht gültig: {line}")

# Spaltenbreite anpassen
for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 30

# Excel-Datei speichern
excel_path = '/Users/thomas/_python/MyApps/Luna_Search/templates/Luna_projects.xlsx'
wb.save(excel_path)
